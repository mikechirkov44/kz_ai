"""Build Excel workbooks for report/catalog exports."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(ws, columns: Sequence[str]) -> None:
    bold = Font(bold=True)
    for idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = bold
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(title) + 2, 12), 36)


def rows_to_workbook(columns: Sequence[str], rows: Iterable[Sequence[Any]], sheet_name: str = "Данные") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _style_header(ws, columns)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def motivation_workbook(report: Any) -> Workbook:
    columns = [
        "Ценовые диапазоны / Номенклатура",
        "ЖЦТ",
        "Дата ЖЦТ",
        "Продано (шт)",
        "Вознаграждение",
        "Итого вознаграждение",
        "Стоимость",
        "Стоимость расчётная",
        "Разница %",
    ]
    rows: list[Sequence[Any]] = []
    groups = list(getattr(report, "groups", None) or [])
    if not groups and getattr(report, "items", None):
        # flat fallback
        for item in report.items:
            rows.append(
                (
                    f"{item.article} {item.name or ''}".strip(),
                    item.lts,
                    item.lts_date,
                    float(item.quantity),
                    float(item.bonus_per_unit),
                    float(item.total_bonus),
                    float(item.cost_amount or 0),
                    float(item.calculated_amount) if item.calculated_amount is not None else None,
                    float(item.difference_percent) if item.difference_percent is not None else None,
                )
            )
    else:
        for group in groups:
            rows.append(
                (
                    f"{group.grade} · {float(group.bonus_per_unit):.0f}",
                    None,
                    None,
                    float(group.quantity),
                    float(group.bonus_per_unit),
                    float(group.total_bonus),
                    float(group.total_cost),
                    float(group.total_calculated_cost) if group.total_calculated_cost else None,
                    float(group.difference_percent) if group.difference_percent is not None else None,
                )
            )
            for item in group.items:
                rows.append(
                    (
                        f"  {item.article} {item.name or ''}".strip(),
                        item.lts,
                        item.lts_date,
                        float(item.quantity),
                        float(item.bonus_per_unit),
                        float(item.total_bonus),
                        float(item.cost_amount or 0),
                        float(item.calculated_amount) if item.calculated_amount is not None else None,
                        float(item.difference_percent) if item.difference_percent is not None else None,
                    )
                )
    rows.append(
        (
            "Итого",
            None,
            None,
            None,
            None,
            float(report.total_bonus),
            float(getattr(report, "total_cost", 0) or 0),
            float(getattr(report, "total_calculated_cost", 0) or 0) or None,
            float(report.difference_percent) if getattr(report, "difference_percent", None) is not None else None,
        )
    )
    wb = rows_to_workbook(columns, rows, "Мотивация")
    meta = wb.create_sheet("Итог", 0)
    meta["A1"] = "Контрагент"
    meta["B1"] = report.counterparty
    meta["A2"] = "Период"
    meta["B2"] = report.period
    meta["A3"] = "Итого вознаграждение"
    meta["B3"] = float(report.total_bonus)
    meta["A4"] = "Стоимость"
    meta["B4"] = float(getattr(report, "total_cost", 0) or 0)
    meta["A5"] = "Стоимость расчётная"
    meta["B5"] = float(getattr(report, "total_calculated_cost", 0) or 0)
    meta["A6"] = "Разница %"
    meta["B6"] = float(report.difference_percent) if getattr(report, "difference_percent", None) is not None else None
    if getattr(report, "clients", None):
        clients_ws = wb.create_sheet("По клиентам")
        _style_header(
            clients_ws,
            ["Контрагент", "Продано", "Строк", "Вознаграждение", "Стоимость", "Расчётная", "Разница %"],
        )
        for idx, row in enumerate(report.clients, start=2):
            clients_ws.cell(row=idx, column=1, value=row.counterparty)
            clients_ws.cell(row=idx, column=2, value=float(row.quantity))
            clients_ws.cell(row=idx, column=3, value=row.lines)
            clients_ws.cell(row=idx, column=4, value=float(row.total_bonus))
            clients_ws.cell(row=idx, column=5, value=float(getattr(row, "total_cost", 0) or 0))
            clients_ws.cell(row=idx, column=6, value=float(getattr(row, "total_calculated_cost", 0) or 0))
            diff = getattr(row, "difference_percent", None)
            clients_ws.cell(row=idx, column=7, value=float(diff) if diff is not None else None)
    return wb


def turnover_matrix_workbook(report: dict) -> Workbook:
    months: list[str] = list(report.get("months") or [])
    view = report.get("view") or "matrix"
    is_main = view == "main"
    base_cols = ["Измерение"]
    if is_main:
        base_cols += ["Артикул", "Тип изделия", "Цвет металла", "ЖЦТ", "Тип работы", "% типа работы"]
    month_cols: list[str] = []
    for m in months:
        if is_main:
            month_cols += [f"{m} Ост.нач", f"{m} Реал.", f"{m} Возвр.", f"{m} Ост.кон", f"{m} Прод."]
        else:
            month_cols += [f"{m} Ост.нач", f"{m} Ост.кон", f"{m} Прод."]
    columns = base_cols + month_cols
    rows: list[list[Any]] = []
    for r in report.get("rows") or []:
        row: list[Any] = [r.get("dimension") or r.get("counterparty") or ""]
        if is_main:
            row += [
                r.get("article"),
                r.get("wear_type"),
                r.get("metal_color"),
                r.get("lts"),
                r.get("work_type"),
                r.get("work_type_percent"),
            ]
        for m in months:
            cell = (r.get("months") or {}).get(m) or {}
            if is_main:
                row += [
                    cell.get("stock_begin", 0),
                    cell.get("realization", 0),
                    cell.get("return_qty", 0),
                    cell.get("stock_end", 0),
                    cell.get("sales", 0),
                ]
            else:
                row += [cell.get("stock_begin", 0), cell.get("stock_end", 0), cell.get("sales", 0)]
        rows.append(row)
    return rows_to_workbook(columns, rows, "Оборачиваемость")


def quarterly_plans_workbook(report: Any) -> Workbook:
    columns = [
        "Головной контрагент",
        "Менеджер",
        "Тип работы",
        "% типа работы",
        "План на квартал",
        "Факт квартал",
        "% выполнения",
        "Динамика",
    ]
    rows = [
        (
            c.counterparty,
            getattr(c, "manager_name", None),
            getattr(c, "work_type_label", None) or getattr(c, "work_type", None),
            float(c.work_type_percent) if getattr(c, "work_type_percent", None) is not None else None,
            float(c.plan),
            float(c.fact),
            float(c.percent),
            c.dynamics,
        )
        for c in report.clients
    ]
    return rows_to_workbook(columns, rows, "ПланФакт")


def nomenclature_workbook(items: Iterable[dict]) -> Workbook:
    columns = ["Артикул", "Наименование", "ЖЦТ", "Дата ЖЦТ", "Тип", "Цвет", "База", "Штрихкод"]
    rows = [
        (
            n.get("article"),
            n.get("name"),
            n.get("lts"),
            n.get("lts_date"),
            n.get("wear_type"),
            n.get("metal_color"),
            n.get("source_id"),
            n.get("barcode"),
        )
        for n in items
    ]
    return rows_to_workbook(columns, rows, "Номенклатура")


def counterparties_workbook(items: Iterable[dict]) -> Workbook:
    columns = ["Наименование", "Тип работы", "%", "Акция", "Менеджер", "Регион", "База", "Магазины"]
    rows = [
        (
            c.get("name"),
            c.get("work_type"),
            c.get("work_type_percent"),
            "да" if c.get("is_promo") else "нет",
            c.get("manager_name"),
            c.get("region"),
            c.get("source_id"),
            ", ".join(c.get("shops") or []),
        )
        for c in items
    ]
    return rows_to_workbook(columns, rows, "Контрагенты")


def documents_workbook(items: Iterable[dict]) -> Workbook:
    columns = ["Дата", "Номер", "Контрагент", "Строк", "Кол-во", "Сумма", "База"]
    rows = [
        (
            d.get("doc_date"),
            d.get("doc_number"),
            d.get("counterparty"),
            d.get("lines"),
            d.get("quantity"),
            d.get("amount"),
            d.get("source_id"),
        )
        for d in items
    ]
    return rows_to_workbook(columns, rows, "Документы")


_BLOCK_FILL = {
    "metal_color": PatternFill("solid", fgColor="CCFBF1"),
    "lts": PatternFill("solid", fgColor="FEF3C7"),
    "wear_type": PatternFill("solid", fgColor="E0E7FF"),
}


def quarterly_summary_workbook(report: Any) -> Workbook:
    """Широкая матрица как на листе 6 ТЗ."""
    labels = report.get("labels") if isinstance(report, dict) else getattr(report, "labels", {}) or {}
    clients = report.get("clients") if isinstance(report, dict) else getattr(report, "clients", []) or []
    block_keys = ("metal_color", "lts", "wear_type")
    block_titles = {
        "metal_color": "Цвет металла",
        "lts": "ЖЦТ",
        "wear_type": "Тип изделия",
    }
    metric_titles = [
        "",
        "ср остаток на квартал",
        labels.get("sales", "итого продажи"),
        labels.get("turnover", "Об-ть квартала"),
        labels.get("avg_turnover", "Ср. об-ть за квартал"),
    ]
    left = ["Контрагент", "Тип работы контрагента", "% типа работ", labels.get("plan", "План отгрузки")]
    right = [
        labels.get("sales_prev", "итого продажи пред. кв."),
        labels.get("sales_prev2", "итого продажи предпред. кв."),
        labels.get("dynamics", "Динамика"),
        "Комментарий",
        labels.get("next_plan", "План работы на след. кв (шт)"),
        "Рекомендации",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Итоги квартала"
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, horizontal="center", vertical="center")
    left_align = Alignment(vertical="center")

    row1 = left + [block_titles[k] for k in block_keys for _ in range(5)] + right
    row2 = [""] * 4 + metric_titles * 3 + [""] * 6
    for idx, value in enumerate(row1, start=1):
        cell = ws.cell(row=1, column=idx, value=value)
        cell.font = bold
        cell.alignment = wrap
    for idx, value in enumerate(row2, start=1):
        cell = ws.cell(row=2, column=idx, value=value)
        cell.font = bold
        cell.alignment = wrap

    # merge left headers vertically, block titles, right headers
    for col in range(1, 5):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    start = 5
    for key in block_keys:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 4)
        fill = _BLOCK_FILL[key]
        for col in range(start, start + 5):
            ws.cell(row=1, column=col).fill = fill
            ws.cell(row=2, column=col).fill = fill
        start += 5
    for col in range(start, start + 6):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    r_idx = 3
    for client in clients:
        matrix = client.get("matrix") or []
        n_rows = max(len(matrix), 1)
        first = r_idx
        last = r_idx + n_rows - 1
        for i, mrow in enumerate(matrix or [{}]):
            is_total = bool(mrow.get("is_total"))
            values: list[Any] = [
                client.get("counterparty") if i == 0 else "",
                client.get("work_type_label") if i == 0 else "",
                client.get("work_type_percent") if i == 0 else "",
                client.get("plan") if i == 0 else "",
            ]
            for key in block_keys:
                cell = mrow.get(key) or {}
                values.extend(
                    [
                        cell.get("dimension"),
                        cell.get("avg_stock"),
                        cell.get("sales_total"),
                        cell.get("quarter_turnover_percent"),
                        cell.get("avg_month_turnover_percent"),
                    ]
                )
            if is_total:
                values.extend(
                    [
                        client.get("sales_prev_quarter"),
                        client.get("sales_prev2_quarter"),
                        client.get("dynamics_percent"),
                        client.get("comment"),
                        client.get("next_quarter_plan"),
                        client.get("recommendations_text"),
                    ]
                )
            else:
                values.extend([None] * 6)
            for c_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = left_align
                if is_total:
                    cell.font = bold
            r_idx += 1
        if last > first:
            for col in (1, 2, 3, 4):
                ws.merge_cells(start_row=first, start_column=col, end_row=last, end_column=col)
        r_idx += 1  # spacer between clients

    widths = [22, 18, 12, 16] + [14] * 15 + [14, 14, 14, 28, 16, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "E3"
    return wb
