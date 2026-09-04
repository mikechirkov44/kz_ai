from decimal import Decimal

from app.middleware_rate_limit import SlidingWindowLimiter
from app.schemas import MotivationClientRow, MotivationItem, MotivationReport
from app.services.export_xlsx import motivation_workbook, workbook_bytes
from fastapi import HTTPException
import pytest


def test_motivation_workbook_bytes():
    report = MotivationReport(
        counterparty="Тест",
        period="2023-01",
        total_bonus=Decimal("6000"),
        items=[
            MotivationItem(
                article="A1",
                name="Кольцо",
                price=Decimal("100"),
                quantity=Decimal("2"),
                grade="1 — 100 000",
                bonus_per_unit=Decimal("50"),
                total_bonus=Decimal("100"),
                lts="Хит",
                lts_date="2022-01-01",
                cost_amount=Decimal("200"),
            )
        ],
        groups=[],
    )
    data = workbook_bytes(motivation_workbook(report))
    assert data[:2] == b"PK"
    assert len(data) > 100


def test_motivation_workbook_all_clients():
    from uuid import uuid4
    from io import BytesIO
    from openpyxl import load_workbook

    cid = uuid4()
    report = MotivationReport(
        counterparty="Все",
        period="2023-01",
        total_bonus=Decimal("1500"),
        clients=[
            MotivationClientRow(
                counterparty_id=cid,
                counterparty="ИП Saona",
                quantity=Decimal("2"),
                lines=1,
                total_bonus=Decimal("1500"),
            )
        ],
        items=[
            MotivationItem(
                article="A1",
                name="Кольцо",
                price=Decimal("95000"),
                quantity=Decimal("2"),
                grade="1 — 100 000",
                bonus_per_unit=Decimal("1500"),
                total_bonus=Decimal("3000"),
                counterparty="ИП Saona",
                counterparty_id=cid,
                cost_amount=Decimal("190000"),
            )
        ],
    )
    wb = load_workbook(BytesIO(workbook_bytes(motivation_workbook(report))))
    assert "По клиентам" in wb.sheetnames
    assert wb["По клиентам"]["A2"].value == "ИП Saona"
    assert wb["Мотивация"]["A1"].value == "Ценовые диапазоны / Номенклатура"


def test_rate_limiter_blocks():
    lim = SlidingWindowLimiter()
    for _ in range(3):
        lim.check("t1", limit=3, window_sec=60)
    with pytest.raises(HTTPException) as exc:
        lim.check("t1", limit=3, window_sec=60)
    assert exc.value.status_code == 429
