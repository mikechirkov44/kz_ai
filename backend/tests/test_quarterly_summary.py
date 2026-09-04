from decimal import Decimal

from app.domain.articles import index_nomenclature, lookup_nomenclature
from app.domain.motivation import work_type_label
from app.domain.quarterly import dim_metrics, recommendations_digest, zip_block_rows
from app.domain.turnover import (
    month_avg_stock,
    quarter_avg_stock,
    sales_dynamics_percent,
    shift_quarter,
)
from app.services.export_xlsx import quarterly_summary_workbook, workbook_bytes


def test_month_and_quarter_avg_stock():
    assert month_avg_stock(Decimal(10), Decimal(20)) == Decimal(15)
    assert quarter_avg_stock([Decimal(15), Decimal(20), Decimal(15)]) == Decimal("50") / Decimal(3)


def test_shift_quarter_wraps_year():
    assert shift_quarter(2026, 1, -1) == (2025, 4)
    assert shift_quarter(2026, 1, -2) == (2025, 3)
    assert shift_quarter(2026, 3, 1) == (2026, 4)
    assert shift_quarter(2026, 4, 1) == (2027, 1)


def test_sales_dynamics_percent():
    assert sales_dynamics_percent(Decimal(43), Decimal(100)) == Decimal(43)
    assert sales_dynamics_percent(Decimal(10), Decimal(0)) is None


def test_dim_metrics_pads_short_month_lists():
    m = dim_metrics(Decimal(10), [Decimal(4)], [Decimal(6)])
    # one month avg 5, two months padded 0 → quarter avg 5/3
    assert m["avg_stock"] == Decimal(5) / Decimal(3)
    # begin/end 10/10 three months, sales 30 → avg stock 10, об-ть 300%, ср.об-ть 100%
    m = dim_metrics(Decimal(30), [Decimal(10)] * 3, [Decimal(10)] * 3)
    assert m["avg_stock"] == Decimal(10)
    assert m["sales_total"] == Decimal(30)
    assert m["quarter_turnover_percent"] == Decimal(300)
    assert m["avg_month_turnover_percent"] == Decimal(100)


def test_zip_block_rows_pads_short_block():
    a = [{"dimension": "Белое 585"}]
    b = [{"dimension": "Актив"}, {"dimension": "Новинка"}]
    c = []
    rows = zip_block_rows(a, b, c)
    assert len(rows) == 2
    assert rows[0][0]["dimension"] == "Белое 585"
    assert rows[1][0] is None
    assert rows[1][1]["dimension"] == "Новинка"
    assert rows[0][2] is None


def test_recommendations_digest_limit():
    items = [{"message": "A"}, {"message": "B"}, {"message": " "}, {"message": "C"}]
    assert recommendations_digest(items, limit=2) == "A B"


def test_work_type_label_ru():
    assert work_type_label("hold") == "Удержание"
    assert work_type_label("Прирост") == "Рост"
    assert work_type_label("падение") == "Падение"
    assert work_type_label(None) == "—"


def test_index_and_lookup_nomenclature():
    class Nom:
        def __init__(self, article, barcode):
            self.article = article
            self.barcode = barcode

    items = [Nom("  00012 ", "B-1")]
    index = index_nomenclature(items)  # type: ignore[arg-type]
    assert lookup_nomenclature(index, "12") is items[0]
    assert lookup_nomenclature(index, "B-1") is items[0]
    assert lookup_nomenclature(index, "missing") is None


def test_quarterly_summary_workbook_matrix():
    report = {
        "labels": {
            "plan": "План отгрузки на 3 квартал",
            "sales": "итого продажи 3 кв",
            "turnover": "Об-ть 3 кв",
            "avg_turnover": "Ср. об-ть за 3 кв",
            "sales_prev": "итого продажи 2 кв.",
            "sales_prev2": "итого продажи 1 кв.",
            "dynamics": "Динамика 3 кв. / 2 кв.",
            "next_plan": "План работы на 4 кв (шт)",
        },
        "clients": [
            {
                "counterparty": "ИП Garant.S",
                "work_type_label": "Удержание",
                "work_type_percent": 0,
                "plan": 50,
                "sales_prev_quarter": 80,
                "sales_prev2_quarter": 70,
                "dynamics_percent": 43,
                "comment": "Участвует в повышенной мотивации",
                "next_quarter_plan": 34,
                "recommendations_text": "Подсортировать кольца Актив Ядро.",
                "matrix": [
                    {
                        "metal_color": {
                            "dimension": "Красное 585",
                            "avg_stock": 10,
                            "sales_total": 8,
                            "quarter_turnover_percent": 80,
                            "avg_month_turnover_percent": 26.67,
                        },
                        "lts": {
                            "dimension": "Актив",
                            "avg_stock": 12,
                            "sales_total": 9,
                            "quarter_turnover_percent": 75,
                            "avg_month_turnover_percent": 25,
                        },
                        "wear_type": {
                            "dimension": "Кольцо",
                            "avg_stock": 5,
                            "sales_total": 4,
                            "quarter_turnover_percent": 80,
                            "avg_month_turnover_percent": 26.67,
                        },
                    },
                    {
                        "is_total": True,
                        "metal_color": {
                            "dimension": "Итого",
                            "avg_stock": 10,
                            "sales_total": 34,
                            "quarter_turnover_percent": 340,
                            "avg_month_turnover_percent": 113.33,
                        },
                        "lts": {
                            "dimension": "Итого",
                            "avg_stock": 10,
                            "sales_total": 34,
                            "quarter_turnover_percent": 340,
                            "avg_month_turnover_percent": 113.33,
                        },
                        "wear_type": {
                            "dimension": "Итого",
                            "avg_stock": 10,
                            "sales_total": 34,
                            "quarter_turnover_percent": 340,
                            "avg_month_turnover_percent": 113.33,
                        },
                    },
                ],
            }
        ],
    }
    data = workbook_bytes(quarterly_summary_workbook(report))
    assert data[:2] == b"PK"
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "Контрагент"
    assert "Цвет металла" in str(ws["E1"].value)
    assert ws["A3"].value == "ИП Garant.S"
    assert ws["E3"].value == "Красное 585"
    assert any(c.value == "Итого" for row in ws.iter_rows(min_row=3, max_row=6, min_col=5, max_col=5) for c in row)
